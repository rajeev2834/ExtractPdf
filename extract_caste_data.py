import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#my File paths
# -----------------------------
INPUT_EXCEL = "C:/Users/NIC/OneDrive/Desktop/CasteCensusData/caste.xlsx"
OUTPUT_DIR = "C:/Users/NIC/OneDrive/Desktop/CasteCensusData/Kashichak/"

COLUMN_MAP = {
    "हाउसहोल्ड फॉर्म सं०": "HouseholdID",
    "जिला": "DistrictName",
    "प्रखंड": "BlockName",
    "पंचायत": "PanchayatName",
    "वार्ड": "WardCode",
    "हाउस न०": "HouseNumber",
    "परिवार के मुखिया का नाम": "FamilyHead",
    "पूरा नाम": "FullName",
    "पिता का नाम": "FatherName",
    "HusbandName": "HusbandName",
    "उम्र": "AgeInYears",
    "लिंग": "GenderCode",
    "Eligible": "Eligible",
    "Scheme Code": "SchemeCode",
    "Remarks": "Remarks"
}

OUTPUT_HINDI_HEADERS = {
    "SLNO": "क्रम संख्या",
    "HouseholdID": "हाउसहोल्ड फॉर्म सं०",
    "WardCode": "वार्ड",
    "FamilyHead": "परिवार के मुखिया का नाम",
    "FullName": "पूरा नाम",
    "FatherName": "पिता का नाम",
    "AgeInYears": "उम्र",
    "GenderCode": "लिंग",
    "Eligible": "Eligible (Y/N)",
    "SchemeCode": "SchemeCode",
    "Remarks": "Remarks"
}

# -----------------------------
# LOAD DATA
# -----------------------------
df = pd.read_excel(INPUT_EXCEL)
df.rename(columns=COLUMN_MAP, inplace=True)

# -----------------------------
# USER INPUT
# -----------------------------
block_name = input("Enter Block Name: ").strip()
panchayat_input = input("Enter Panchayat Name (Enter for ALL): ").strip()
ward_input = input("Enter Ward No (Enter for ALL): ").strip()

# -----------------------------
# FILTER BLOCK
# -----------------------------
block_df = df[df["BlockName"].str.lower() == block_name.lower()]
if block_df.empty:
    raise ValueError("No data found for this block")

district = block_df["DistrictName"].iloc[0]

if panchayat_input:
    block_df = block_df[block_df["PanchayatName"].str.lower() == panchayat_input.lower()]

if ward_input:
    block_df = block_df[block_df["WardCode"].astype(str) == ward_input]

# -----------------------------
# PANCHAYAT-WISE OUTPUT
# -----------------------------
for panchayat, panchayat_df in block_df.groupby("PanchayatName"):
    
    eldest_df = panchayat_df.loc[
        panchayat_df.groupby("HouseholdID")["AgeInYears"].idxmax()
    ]

    eldest_df = eldest_df.drop(
        columns=[
            "DistrictName",
            "BlockName",
            "HouseNumber",
            "HusbandName",
            "PanchayatName"   # ❌ remove panchayat from output
        ],
        errors="ignore"
    )


    eldest_df.sort_values(by=["WardCode", "HouseholdID"], inplace=True)
    eldest_df.insert(0, "SLNO", range(1, len(eldest_df) + 1))
    eldest_df.rename(columns=OUTPUT_HINDI_HEADERS, inplace=True)
    eldest_df.reset_index(drop=True, inplace=True)

    output_file = f"{OUTPUT_DIR}{block_name}_{panchayat.replace(' ', '_')}.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        eldest_df.to_excel(writer, index=False, startrow=2)

    wb = load_workbook(output_file)
    ws = wb.active

    # -------- Title Row --------
    header_text = f"जिला : {district}, प्रखंड : {block_name}, पंचायत : {panchayat}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    hcell = ws.cell(row=1, column=1)
    hcell.value = header_text
    hcell.font = Font(bold=True, size=16)
    hcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # -------- Footer (Repeat on ALL Pages) --------
    ws.oddFooter.left.text = f"1-MMRY, 2-SJY, 3-MMUY, 4-MMLUY, 5-PMAY, 6-Pension, 7-RC, 8-PM Vishwakarma, 9-PM Kisan, 10-Others"
    ws.oddFooter.right.text = "Page &P of &N"

    # Same footer for even pages (important!)
    ws.oddFooter.left.text = f"1-MMRY, 2-SJY, 3-MMUY, 4-MMLUY, 5-PMAY, 6-Pension, 7-RC, 8-PM Vishwakarma, 9-PM Kisan, 10-Others"
    ws.oddFooter.right.text = "Page &P of &N"


    # -------- Column Header Row Formatting (Row 3) --------
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        # Add border to header row
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set header row height
    ws.row_dimensions[3].height = 33

    # -------- Data Rows Formatting --------
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting to all data rows (starting from row 4)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        # Set row height for all data rows
        ws.row_dimensions[row[0].row].height = 33
        
        for cell in row:
            # Apply text wrapping and center alignment
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            # Apply borders
            cell.border = thin_border
            
            # Apply font for data rows
            cell.font = Font(size=11)

    # -------- Page Setup (EXACT specifications from screenshot) --------
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1  # Fit to 1 page wide
    
    # Set margins (IN INCHES - exact values from your screenshot)
    ws.page_margins.left = 0.25    # 0.25 inches
    ws.page_margins.right = 0.08   # 0.08 inches
    ws.page_margins.top = 0.10     # 0.10 inches
    ws.page_margins.bottom = 0.157 # 0.157 inches (from screenshot)
    ws.page_margins.header = 0.16  # 0.16 inches
    ws.page_margins.footer = 0.157 # 0.157 inches (from screenshot)

    # -------- Set Repeating Headers on Each Page --------
    # This makes row 3 (column headers) repeat on every page
    ws.print_title_rows = "1:3"  # Rows 1 to 3 (Title + column headers)

    # -------- Column Widths --------
    # Set fixed uniform column widths for better printing
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # -------- Set Print Area --------
    # Include all rows and columns in print area
    ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # -------- Freeze Panes (optional) --------
    # Freeze the title and header rows
    ws.freeze_panes = "A4"

    # Save the workbook
    wb.save(output_file)

    print(f"✔ Printable Excel generated: {output_file}")